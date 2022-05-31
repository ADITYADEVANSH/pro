from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from win32com import client
import os

def mainCode ():
    
    files=[]

    temp_data=[]

    try:
       # tempath = 'C:/Users/user/Documents/Python/excel project/Demo/new certificates'
        #C:\Users\user\Documents\Python\excel project\Demo\new certificates
        path = input("Enter path for xlsx files : ")
        path = path.replace("\\", "/")
        tempath = input("Enter path for template file : ")
        tempath = tempath.replace("\\", "/")

        # path = 'C:/Users/user/Documents/Python/excel project/Demo/xlsx files'
        #C:\Users\user\Documents\Python\excel project\Demo\xlsx files
        try:
            twb = load_workbook(tempath + '/template.xlsx')     # when name of all the templates is same i.e. 'template'
            tws = twb.active
        except:
            print("Make sure the path is correct and name of the template is \"template.xlsx\"")

        for row in range(21,49):
            temp_row=[]
            
            for col in range(1,19):
                temp_row.append(tws[ chr(64+col) + str(row) ].value)
                
            temp_data.append(temp_row)
               
        twb.save(tempath + '/template.xlsx')        

        for x in os.listdir(path) :
            if '.xlsx' in x :
        #        files.append(x)
                 wb = load_workbook(path + '/' + x)
                 ws = wb.active


                 for row in range(21,49):
                     for col in range(1,19):
                         try:
                             ws[ chr(64+col) + str(row) ] = temp_data[row-21][col-1]
                         except AttributeError:
                             continue
                            
                 img = Image('C:/Users/user/Documents/Python/sign.png')
                 img.width = 124.72
                 img.height = 124.72
                 ws.add_image(img,'A67')                  
                 wb.save(tempath + '/' + x)
            else:
                print("      NO XLSX FILES IN FOLDER :(, FIRST CONVERT XLS FILES TO XLSX FILES")
    except:
        print("An exception occured")
    else:
        print("--------------------Created new certificates---------------------")


def xlsToXlsx() :
    excel = client.Dispatch("Excel.Application")
    
    #path = 'C:/Users/user/Documents/Python/excel project/Demo/xls files'               #xls file location
    #path2 = 'C:\\Users\\user\\Documents\\Python\\excel project\\Demo\\xls files'           #same path but with double slashes
    path = input("Enter path for xls files : ")
    path2 = input("Confirm path : ")
    #path2 = path[::-1][::-1]
    #print(path2)
    path2 = path2.replace("\\", "\\")
    path = path.replace("\\", "/")
    #print(path + "\n" + path2)
    try:
        for file in os.listdir(path):
            filename, fileextention = os.path.splitext(file)
            wb = excel.Workbooks.Open(path + '/' + file)
            output_path = path2 + '\\' + filename + '.xlsx'
            wb.SaveAs(output_path,51)
            wb.close
            os.remove(path + '/' + file)  
        excel.quit()
    except:
        print("An exception occurred, clear excel from task manager and make sure path is correct.")
    else:
        print("----------------------Converted xls to xlsx----------------------")
        
choice=0
print("|---------------------------------------------------------------|")
print("|                     AUTOMATE MY EXCEL                         |")
print("|---------------------------------------------------------------|")

while(choice!='3'):
    print("-"*65)
    print("\t\t 1: Convert xls files to xlsx files")
    print("\t\t 2: Create new certificates")
    print("\t\t 3: EXIT")
    choice=input("\t\t Enter your choice : ")
    if choice=='1':
        xlsToXlsx()
    elif choice=='2':
        mainCode()
    elif choice!='1' and choice!='2' and choice!='3':
        print("Wrong input; input numbers from 1 to 3")
    
else :
    print("\t\t-->Exiting loop")

